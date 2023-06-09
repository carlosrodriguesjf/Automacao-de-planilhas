# PROJETO 2 - MESTRES DA AUTOMAÇÃO - AUTOMAÇÃO DE PLANILHAS

import openpyxl

# Criação do workbook e da Sheet 
workbook = openpyxl.Workbook()


# Declaração de variáveis
criar_pagina = 's'
adiciona_coluna = 's'
adiciona_dados = 's'
colunas = []
dados =[]

# Tela de boas vindas e criação das planilhas do workbook
print('\nBem-vindo ao gerador de planilhas!\nPara começar vamos criar uma nova página dentro de uma planilha')

# Criação das plnilhas
while criar_pagina == 's':
    nome_pagina = input('Digite o nome da página: ')
    workbook.create_sheet(nome_pagina)
    criar_pagina = input('Criar mais uma página nesta planilha?(s/n): ')
del workbook['Sheet']
print(workbook.sheetnames)    


pagina_manipulada = input('Digite o nome da página a ser manipulada: ')
sheet_atual = workbook[pagina_manipulada]

# Inserção do cabeçalho das planilhas
while adiciona_coluna == 's':
    nomes = input('Digite uma coluna para o cabeçalho: ')
    colunas.append(nomes)  
    adiciona_coluna =  input('Adicionar mais uma coluna?(s/n): ')
sheet_atual.append(colunas)

# Adição de dados
adiciona_dados = input('Adicionar dados a essa planilha?(s/n): ')
print('As páginas disponíveis são: ',workbook.sheetnames)
pagina_dados = input('Em qual página devemos adicionar dados?: ')
sheet_atual = workbook[pagina_dados]

while adiciona_dados == 's':
    colunas_dados = input('Digite os dados a serem adicionados a uma nova linha, separados por vírgula: ')
    lista_coluna_dados = colunas_dados.split(',')
    sheet_atual.append(lista_coluna_dados)
    adiciona_dados = input('Adicionar nova linha? (s/n) ')

# Savamento da planilha
nome_planilha = input('DIgite o nome da planilha a ser salva: ')
workbook.save(nome_planilha + '.xlsx')

print('Planilha salva com sucesso')